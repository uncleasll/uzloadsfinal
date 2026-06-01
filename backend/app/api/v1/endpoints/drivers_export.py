"""
Driver list export — PDF and Excel.
GET /api/v1/drivers/export/pdf
GET /api/v1/drivers/export/xlsx
"""
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import date
import io

from app.db.session import get_db
from app.models.models import Driver, DriverProfile, Truck, Trailer
from app.services.company_service import company_identity_lines, get_company, resolve_logo_file

router = APIRouter(tags=["drivers-export"])


def _fmt_date(v) -> str:
    if not v:
        return ""
    try:
        if hasattr(v, "strftime"):
            return v.strftime("%m/%d/%y")
    except Exception:
        pass
    parts = str(v).split("-")
    if len(parts) == 3:
        return f"{parts[1]}/{parts[2]}/{parts[0][2:]}"
    return str(v)


def _get_driver_rows(db: Session):
    """Return list of dicts for all active drivers."""
    drivers = db.query(Driver).filter(Driver.is_active == True).order_by(Driver.name).all()
    rows = []
    for d in drivers:
        p = db.query(DriverProfile).filter(DriverProfile.driver_id == d.id).first()
        truck = db.query(Truck).filter(Truck.id == p.truck_id).first() if p and p.truck_id else None
        trailer = db.query(Trailer).filter(Trailer.id == p.trailer_id).first() if p and p.trailer_id else None
        rows.append({
            "name":       f"{d.name} [{d.driver_type}]",
            "type":       d.driver_type or "",
            "status":     (p.driver_status if p else "") or "",
            "hire_date":  _fmt_date(p.hire_date if p else None),
            "term_date":  _fmt_date(p.termination_date if p else None),
            "phone":      d.phone or "",
            "email":      d.email or "",
            "truck":      truck.unit_number if truck else "",
            "trailer":    trailer.unit_number if trailer else "",
            "payable_to": (p.payable_to if p and p.payable_to else d.name) or "",
        })
    return rows


# ── PDF ───────────────────────────────────────────────────────────────────────

@router.get("/drivers/export/pdf")
def export_drivers_pdf(db: Session = Depends(get_db)):
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
    from reportlab.lib.styles import ParagraphStyle

    rows = _get_driver_rows(db)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        rightMargin=0.6 * inch, leftMargin=0.6 * inch,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
    )

    DARK  = colors.HexColor("#111827")
    GRAY  = colors.HexColor("#6b7280")
    TH_BG = colors.HexColor("#1a2332")

    th_s = ParagraphStyle("th", fontName="Helvetica-Bold", fontSize=7, textColor=colors.white)
    td_s = ParagraphStyle("td", fontName="Helvetica",      fontSize=7, textColor=DARK, leading=10)
    hd_s = ParagraphStyle("hd", fontName="Helvetica",      fontSize=9, textColor=DARK, leading=14)
    tt_s = ParagraphStyle("tt", fontName="Helvetica-Bold", fontSize=12, textColor=DARK, spaceAfter=8)

    story = []

    # Company header — pulled live from "My Company" settings
    company = get_company(db)
    logo_file = resolve_logo_file(company.get("logo_path") or "")
    if logo_file:
        try:
            logo = Image(logo_file)
            ratio = min((1.45 * inch) / logo.imageWidth, (0.55 * inch) / logo.imageHeight)
            logo.drawWidth = logo.imageWidth * ratio
            logo.drawHeight = logo.imageHeight * ratio
        except Exception:
            logo = Paragraph(f"<b>{company.get('name') or 'My Company'}</b>", hd_s)
    else:
        logo = Paragraph(f"<b>{company.get('name') or 'My Company'}</b>", hd_s)
    header_html = "<br/>".join(
        f"<b>{line}</b>" if i == 0 else line
        for i, line in enumerate(company_identity_lines(company))
    )
    header = Table([[logo, Paragraph(header_html, hd_s)]], colWidths=[2.0 * inch, 4.9 * inch])
    header.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(header)
    story.append(Spacer(1, 0.2 * inch))
    story.append(Paragraph("<b>Drivers</b>", tt_s))

    # Table columns
    COLS = ["Name", "Type", "Status", "Hire date", "Term date", "Phone", "Email", "Truck", "Trailer", "Payable to"]
    COL_W = [1.4*inch, 0.45*inch, 0.55*inch, 0.65*inch, 0.65*inch,
             0.85*inch, 1.5*inch, 0.55*inch, 0.55*inch, 1.0*inch]

    table_data = [[Paragraph(c, th_s) for c in COLS]]
    for r in rows:
        table_data.append([
            Paragraph(r["name"],       td_s),
            Paragraph(r["type"],       td_s),
            Paragraph(r["status"],     td_s),
            Paragraph(r["hire_date"],  td_s),
            Paragraph(r["term_date"],  td_s),
            Paragraph(r["phone"],      td_s),
            Paragraph(r["email"],      td_s),
            Paragraph(r["truck"],      td_s),
            Paragraph(r["trailer"],    td_s),
            Paragraph(r["payable_to"], td_s),
        ])

    t = Table(table_data, colWidths=COL_W, repeatRows=1)
    t.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), TH_BG),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ("LEFTPADDING",   (0, 0), (-1, -1), 4),
    ]))
    story.append(t)

    def _page_num(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(GRAY)
        canvas.drawCentredString(letter[0] / 2, 0.35 * inch, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=_page_num, onLaterPages=_page_num)
    buf.seek(0)
    fname = f"drivers_{date.today().strftime('%Y%m%d')}.pdf"
    return StreamingResponse(
        buf, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )


# ── Excel ─────────────────────────────────────────────────────────────────────

@router.get("/drivers/export/xlsx")
def export_drivers_xlsx(db: Session = Depends(get_db)):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rows = _get_driver_rows(db)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Drivers"

    hdr_fill = PatternFill("solid", fgColor="1a2332")
    hdr_font = Font(bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers   = ["Name", "Type", "Status", "Hire Date", "Term Date", "Phone", "Email", "Truck", "Trailer", "Payable To"]
    col_widths = [32, 8, 12, 12, 12, 16, 32, 10, 10, 24]

    for i, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=i, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 18

    for ri, r in enumerate(rows, 2):
        fill = PatternFill("solid", fgColor="FFFFFF" if ri % 2 == 0 else "F9FAFB")
        for ci, val in enumerate([
            r["name"], r["type"], r["status"], r["hire_date"], r["term_date"],
            r["phone"], r["email"], r["truck"], r["trailer"], r["payable_to"],
        ], 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = fill
            cell.font = Font(size=9)
            cell.alignment = Alignment(vertical="center")
            cell.border = border
        ws.row_dimensions[ri].height = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"drivers_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"}
    )
