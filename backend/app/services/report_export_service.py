"""
Report generation service — PDF (ReportLab) and Excel (openpyxl).
Matches the exact layout shown in the sample PDF:
  - Company logo area (truck icon placeholder) + bold company name/email/phone top-right
  - Bold report title
  - Bold meta lines (Dates range, Dispatcher, Statuses, Billing statuses, etc.)
  - Clean bordered table with shaded header
  - Total row at the bottom
"""

import io
from datetime import datetime
from typing import List, Dict, Any, Optional

# ── ReportLab ──────────────────────────────────────────────────────────────────
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph,
    Spacer, HRFlowable
)
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

# ── openpyxl ──────────────────────────────────────────────────────────────────
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Colors ────────────────────────────────────────────────────────────────────
BLACK   = colors.HexColor("#000000")
GRAY_BG = colors.HexColor("#f5f5f5")
BORDER  = colors.HexColor("#cccccc")
WHITE   = colors.white

COMPANY_INFO = {
    "name":  "Silkroad llc",
    "email": "Email: asilbekkarimov066@gmail.com",
    "phone": "Phone: (970) 610-8065",
}


# ─────────────────────────────────────────────────────────────────────────────
# Shared helpers
# ─────────────────────────────────────────────────────────────────────────────

def _fmt_currency(v) -> str:
    try:
        return f"${float(v):,.2f}"
    except Exception:
        return str(v)


def _fmt_num(v) -> str:
    try:
        return f"{int(v):,}"
    except Exception:
        return str(v)


def _fmt_date(s) -> str:
    if not s or s == "None":
        return ""
    try:
        d = str(s)
        if "-" in d:
            parts = d.split("-")
            return f"{parts[1]}/{parts[2]}/{parts[0][2:]}"
        return d
    except Exception:
        return str(s)


# ─────────────────────────────────────────────────────────────────────────────
# PDF generation
# ─────────────────────────────────────────────────────────────────────────────

def _pdf_doc(buffer, landscape_mode=False):
    ps = landscape(letter) if landscape_mode else letter
    return SimpleDocTemplate(
        buffer, pagesize=ps,
        rightMargin=0.6*inch, leftMargin=0.6*inch,
        topMargin=0.6*inch, bottomMargin=0.6*inch,
    )


def _header_table(story, report_title: str, meta_lines: List[str]):
    """Build the company header + report title + meta block matching sample PDF."""

    # Company info (right side)
    company_text = (
        f"<b>{COMPANY_INFO['name']}</b><br/>"
        f"{COMPANY_INFO['email']}<br/>"
        f"{COMPANY_INFO['phone']}"
    )

    # Logo placeholder (simple truck icon text, since we can't embed the image)
    logo_style = ParagraphStyle("logo", fontSize=9, leading=11, textColor=BLACK)
    company_style = ParagraphStyle("company", fontSize=9, leading=13, textColor=BLACK)
    title_style   = ParagraphStyle("title",   fontSize=12, leading=15, textColor=BLACK, fontName="Helvetica-Bold", spaceBefore=6)
    meta_style    = ParagraphStyle("meta",    fontSize=9,  leading=12, textColor=BLACK)
    meta_bold     = ParagraphStyle("metab",   fontSize=9,  leading=12, textColor=BLACK, fontName="Helvetica-Bold")

    # 2-column header: [logo | company info]
    header_data = [[
        Paragraph("<b>TOPTRUCK</b><br/><font size='7'>COMPANY</font>", logo_style),
        Paragraph(company_text, company_style),
    ]]
    header_tbl = Table(header_data, colWidths=["50%", "50%"])
    header_tbl.setStyle(TableStyle([
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("ALIGN", (0,0), (0,0), "LEFT"),
        ("ALIGN", (1,0), (1,0), "RIGHT"),
        ("BOTTOMPADDING", (0,0), (-1,-1), 8),
    ]))
    story.append(header_tbl)
    story.append(Spacer(1, 10))

    # Report title
    story.append(Paragraph(f"<b>{report_title}</b>", title_style))
    story.append(Spacer(1, 6))

    # Meta lines — bold label, normal value on same line
    for line in meta_lines:
        if ": " in line:
            label, value = line.split(": ", 1)
            story.append(Paragraph(f"<b>{label}:</b> {value}", meta_style))
        else:
            story.append(Paragraph(line, meta_style))
    story.append(Spacer(1, 14))


def _data_table(story, headers: List[str], rows: List[List[str]], totals: Optional[List[str]] = None, col_widths=None, right_align_cols: List[int] = None):
    """Build the main data table matching sample PDF style."""
    right_align_cols = right_align_cols or []

    header_style = ParagraphStyle("th", fontSize=8, leading=10, fontName="Helvetica-Bold", textColor=BLACK)
    cell_style   = ParagraphStyle("td", fontSize=8, leading=10, textColor=BLACK)

    def _cell(text, bold=False, align="LEFT"):
        s = ParagraphStyle("c", fontSize=8, leading=10,
                           textColor=BLACK,
                           fontName="Helvetica-Bold" if bold else "Helvetica",
                           alignment={"LEFT": TA_LEFT, "RIGHT": TA_RIGHT, "CENTER": TA_CENTER}[align])
        return Paragraph(str(text) if text is not None else "", s)

    # Header row
    table_data = [[
        _cell(h, bold=True, align="RIGHT" if i in right_align_cols else "LEFT")
        for i, h in enumerate(headers)
    ]]

    # Data rows
    for row in rows:
        table_data.append([
            _cell(str(v) if v is not None else "", align="RIGHT" if i in right_align_cols else "LEFT")
            for i, v in enumerate(row)
        ])

    # Totals row
    if totals:
        table_data.append([
            _cell(str(v) if v is not None else "", bold=True, align="RIGHT" if i in right_align_cols else "LEFT")
            for i, v in enumerate(totals)
        ])

    n_cols = len(headers)
    available = 7.3 * inch  # letter width minus margins
    if col_widths is None:
        col_widths = [available / n_cols] * n_cols

    tbl = Table(table_data, colWidths=col_widths, repeatRows=1)

    style = [
        # Header
        ("BACKGROUND", (0,0), (-1,0), GRAY_BG),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        # Grid
        ("GRID", (0,0), (-1,-1), 0.5, BORDER),
        ("ROWBACKGROUNDS", (1,1), (-1,-2), [WHITE, GRAY_BG]),
        # Totals row
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
    ]

    if totals:
        style += [
            ("BACKGROUND", (0,-1), (-1,-1), GRAY_BG),
            ("FONTNAME", (0,-1), (-1,-1), "Helvetica-Bold"),
            ("LINEABOVE", (0,-1), (-1,-1), 1, BLACK),
        ]

    tbl.setStyle(TableStyle(style))
    story.append(tbl)


# ─────────────────────────────────────────────────────────────────────────────
# Excel generation
# ─────────────────────────────────────────────────────────────────────────────

def _xlsx_workbook(report_title: str, meta_lines: List[str], headers: List[str],
                   rows: List[List], totals: Optional[List] = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = report_title[:31]

    # ── Styles ──
    header_font  = Font(name="Calibri", bold=True, size=11)
    title_font   = Font(name="Calibri", bold=True, size=13)
    meta_font    = Font(name="Calibri", bold=False, size=10)
    meta_b_font  = Font(name="Calibri", bold=True, size=10)
    cell_font    = Font(name="Calibri", size=10)
    total_font   = Font(name="Calibri", bold=True, size=10)
    company_font = Font(name="Calibri", bold=True, size=11)

    header_fill = PatternFill("solid", fgColor="F2F2F2")
    total_fill  = PatternFill("solid", fgColor="F2F2F2")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    thick_top = Border(left=thin, right=thin, top=Side(style="medium", color="000000"), bottom=thin)

    def _set(row, col, value, font=None, fill=None, align=None, border_=None, num_fmt=None):
        c = ws.cell(row=row, column=col, value=value)
        if font:   c.font = font
        if fill:   c.fill = fill
        if align:  c.alignment = align
        if border_: c.border = border_
        if num_fmt: c.number_format = num_fmt

    row_num = 1

    # Company header
    ws.merge_cells(f"A{row_num}:C{row_num}")
    _set(row_num, 1, "TOPTRUCK COMPANY", font=company_font)
    ws.merge_cells(f"D{row_num}:G{row_num}")
    _set(row_num, 4, COMPANY_INFO["name"], font=company_font,
         align=Alignment(horizontal="right"))
    row_num += 1
    ws.merge_cells(f"D{row_num}:G{row_num}")
    _set(row_num, 4, COMPANY_INFO["email"],
         align=Alignment(horizontal="right"), font=cell_font)
    row_num += 1
    ws.merge_cells(f"D{row_num}:G{row_num}")
    _set(row_num, 4, COMPANY_INFO["phone"],
         align=Alignment(horizontal="right"), font=cell_font)
    row_num += 2

    # Report title
    n_cols = len(headers)
    end_col = get_column_letter(n_cols)
    ws.merge_cells(f"A{row_num}:{end_col}{row_num}")
    _set(row_num, 1, report_title, font=title_font)
    row_num += 1

    # Meta lines
    for line in meta_lines:
        ws.merge_cells(f"A{row_num}:{end_col}{row_num}")
        if ": " in line:
            label, value = line.split(": ", 1)
            _set(row_num, 1, f"{label}: {value}", font=meta_font)
        else:
            _set(row_num, 1, line, font=meta_font)
        row_num += 1
    row_num += 1

    # Column headers
    for ci, h in enumerate(headers, 1):
        _set(row_num, ci, h,
             font=header_font, fill=header_fill, border_=border,
             align=Alignment(horizontal="center", wrap_text=True))
    row_num += 1

    # Data rows
    for ri, row in enumerate(rows):
        bg = PatternFill("solid", fgColor="FFFFFF") if ri % 2 == 0 else PatternFill("solid", fgColor="FAFAFA")
        for ci, val in enumerate(row, 1):
            align = Alignment(horizontal="right") if isinstance(val, (int, float)) else Alignment(horizontal="left")
            _set(row_num, ci, val, font=cell_font, fill=bg, border_=border, align=align)
        row_num += 1

    # Totals
    if totals:
        for ci, val in enumerate(totals, 1):
            align = Alignment(horizontal="right") if isinstance(val, (int, float)) else Alignment(horizontal="left")
            _set(row_num, ci, val, font=total_font, fill=total_fill,
                 border_=thick_top, align=align)
        row_num += 1

    # Auto-width columns
    for ci, h in enumerate(headers, 1):
        col_letter = get_column_letter(ci)
        max_len = max(len(str(h)), *(len(str(r[ci-1])) for r in rows) if rows else [0])
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────────────────────
# Report-specific builders
# ─────────────────────────────────────────────────────────────────────────────

def generate_rate_per_mile_pdf(report_data: dict, filters: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf, landscape_mode=True)
    story = []

    meta = [
        f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}",
        f"Dispatcher: {filters.get('dispatcher_name', 'All dispatchers')}",
        f"Statuses: {filters.get('statuses', 'All')}",
        f"Billing statuses: {filters.get('billing_statuses', 'All')}",
    ]
    _header_table(story, "Rate per Mile", meta)

    headers = ["Pickup date","Completed date","Load #","Truck #","Driver","Dispatcher","Route","Empty miles","Loaded miles","Total miles","Rate","Rate per mile"]
    right_cols = [7,8,9,10,11]

    rows_out = []
    for r in report_data.get("rows", []):
        route = f"{r.get('pickup_city','')}, {r.get('pickup_state','')} - {r.get('delivery_city','')}, {r.get('delivery_state','')}"
        rows_out.append([
            _fmt_date(r.get("pickup_date")), _fmt_date(r.get("actual_delivery_date")),
            r.get("load_number",""), r.get("truck",""),
            r.get("driver",""), r.get("dispatcher",""),
            route,
            _fmt_num(r.get("empty_miles",0)), _fmt_num(r.get("loaded_miles",0)),
            _fmt_num(r.get("total_miles",0)),
            _fmt_currency(r.get("rate",0)), _fmt_currency(r.get("rate_per_mile",0)),
        ])

    s = report_data.get("summary",{})
    total_loaded = sum(r.get("loaded_miles",0) for r in report_data.get("rows",[]))
    total_miles  = sum(r.get("total_miles",0) for r in report_data.get("rows",[]))
    total_rate   = s.get("total_revenue",0)
    avg_rpm      = total_rate / total_miles if total_miles else 0

    totals = ["","","","","","","Total:",
              _fmt_num(0), _fmt_num(total_loaded), _fmt_num(total_miles),
              _fmt_currency(total_rate), _fmt_currency(avg_rpm)]

    avail = 10.1 * inch
    cw = [0.65,0.65,0.5,0.6,1.0,0.8,1.5,0.6,0.6,0.6,0.65,0.65]
    total = sum(cw)
    cw = [x/total*avail for x in cw]

    _data_table(story, headers, rows_out, totals, col_widths=cw, right_align_cols=right_cols)
    doc.build(story)
    buf.seek(0)
    return buf.read()


def generate_rate_per_mile_xlsx(report_data: dict, filters: dict) -> bytes:
    meta = [
        f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}",
        f"Dispatcher: {filters.get('dispatcher_name', 'All dispatchers')}",
        f"Statuses: {filters.get('statuses', 'All')}",
        f"Billing statuses: {filters.get('billing_statuses', 'All')}",
    ]
    headers = ["Pickup date","Completed date","Load #","Truck #","Driver","Dispatcher","Route","Empty miles","Loaded miles","Total miles","Rate","Rate per mile"]
    rows_out = []
    for r in report_data.get("rows", []):
        route = f"{r.get('pickup_city','')}, {r.get('pickup_state','')} - {r.get('delivery_city','')}, {r.get('delivery_state','')}"
        rows_out.append([
            _fmt_date(r.get("pickup_date")), _fmt_date(r.get("actual_delivery_date")),
            r.get("load_number",""), r.get("truck",""), r.get("driver",""), r.get("dispatcher",""),
            route,
            r.get("empty_miles",0), r.get("loaded_miles",0), r.get("total_miles",0),
            r.get("rate",0), r.get("rate_per_mile",0),
        ])
    s = report_data.get("summary",{})
    total_miles = sum(r.get("total_miles",0) for r in report_data.get("rows",[]))
    total_loaded = sum(r.get("loaded_miles",0) for r in report_data.get("rows",[]))
    total_rate = s.get("total_revenue",0)
    avg_rpm = total_rate/total_miles if total_miles else 0
    totals = ["","","","","","","Total:", 0, total_loaded, total_miles, total_rate, avg_rpm]
    return _xlsx_workbook("Rate per Mile", meta, headers, rows_out, totals)


def generate_total_revenue_pdf(report_data: dict, filters: dict, columns: list) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf, landscape_mode=True)
    story = []
    meta = [
        f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}",
        f"Statuses: {filters.get('statuses','All')}",
        f"Billing statuses: {filters.get('billing_statuses','All')}",
    ]
    _header_table(story, "Total Revenue Report", meta)

    COL_MAP = {
        "pickup_date": ("Pickup date", lambda r: _fmt_date(r.get("pickup_date"))),
        "actual_delivery_date": ("Completed date", lambda r: _fmt_date(r.get("actual_delivery_date"))),
        "load_number": ("Load #", lambda r: r.get("load_number","")),
        "route": ("Route", lambda r: f"{r.get('pickup_city','')}, {r.get('pickup_state','')} - {r.get('delivery_city','')}, {r.get('delivery_state','')}"),
        "broker": ("Broker", lambda r: r.get("broker","")),
        "po_number": ("PO #", lambda r: r.get("po_number","")),
        "rate": ("Invoice amount", lambda r: _fmt_currency(r.get("rate",0))),
        "driver": ("Driver", lambda r: r.get("driver","")),
        "truck": ("Truck #", lambda r: r.get("truck","")),
        "driver_pay": ("Driver pay", lambda r: _fmt_currency(r.get("driver_pay",0))),
    }
    active_cols = [k for k in columns if k in COL_MAP]
    if not active_cols:
        active_cols = ["pickup_date","actual_delivery_date","load_number","route","rate"]

    headers = [COL_MAP[k][0] for k in active_cols]
    rows_out = [[COL_MAP[k][1](r) for k in active_cols] for r in report_data.get("rows",[])]

    s = report_data.get("summary",{})
    total_revenue = s.get("total_revenue",0)
    totals = []
    for k in active_cols:
        if k == "rate": totals.append(_fmt_currency(total_revenue))
        elif k == "route": totals.append("Total:")
        elif k == "driver_pay": totals.append(_fmt_currency(s.get("total_driver_pay",0)))
        else: totals.append("")

    right_cols = [i for i, k in enumerate(active_cols) if k in ("rate","driver_pay")]
    avail = 10.1 * inch
    cw = [avail/len(headers)]*len(headers)
    _data_table(story, headers, rows_out, totals if any(t for t in totals) else None, col_widths=cw, right_align_cols=right_cols)
    doc.build(story)
    buf.seek(0)
    return buf.read()


def generate_total_revenue_xlsx(report_data: dict, filters: dict, columns: list) -> bytes:
    meta = [
        f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}",
        f"Statuses: {filters.get('statuses','All')}",
        f"Billing statuses: {filters.get('billing_statuses','All')}",
    ]
    headers = ["Pickup date","Completed date","Load #","Route","Broker","PO #","Invoice amount","Driver","Truck #","Driver pay","Status","Billing"]
    rows_out = []
    for r in report_data.get("rows",[]):
        route = f"{r.get('pickup_city','')}, {r.get('pickup_state','')} - {r.get('delivery_city','')}, {r.get('delivery_state','')}"
        rows_out.append([_fmt_date(r.get("pickup_date")),_fmt_date(r.get("actual_delivery_date")),r.get("load_number",""),route,r.get("broker",""),r.get("po_number",""),r.get("rate",0),r.get("driver",""),r.get("truck",""),r.get("driver_pay",0),r.get("status",""),r.get("billing_status","")])
    s = report_data.get("summary",{})
    totals = ["","","","","","","Total",_fmt_currency(s.get("total_revenue",0)),"","","",""]
    return _xlsx_workbook("Total Revenue Report", meta, headers, rows_out, totals)


def generate_gross_profit_pdf(report_data: dict, filters: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf)
    story = []
    meta = [
        f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}",
    ]
    if report_data.get("driver_name"): meta.append(f"Driver: {report_data['driver_name']}")
    if report_data.get("truck_unit"):  meta.append(f"Truck: {report_data['truck_unit']}")
    _header_table(story, "Gross Profit Report", meta)

    s = report_data.get("summary",{})
    headers = ["", "Amount"]
    rows_out = [
        ["Total Revenue",     _fmt_currency(s.get("total_revenue",0))],
        ["Loads Revenue",     _fmt_currency(s.get("loads_revenue",0))],
        ["Other Revenue",     _fmt_currency(s.get("other_revenue",0))],
        ["Driver Payments",   _fmt_currency(s.get("driver_payments",0))],
        ["Fuel",              _fmt_currency(s.get("fuel",0))],
        ["Tolls",             _fmt_currency(s.get("tolls",0))],
        ["Gross Profit",      _fmt_currency(s.get("gross_profit",0))],
    ]
    cw = [5.0*inch, 2.3*inch]
    _data_table(story, headers, rows_out, col_widths=cw, right_align_cols=[1])
    doc.build(story)
    buf.seek(0)
    return buf.read()


def generate_gross_profit_xlsx(report_data: dict, filters: dict) -> bytes:
    meta = [f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}"]
    if report_data.get("driver_name"): meta.append(f"Driver: {report_data['driver_name']}")
    if report_data.get("truck_unit"):  meta.append(f"Truck: {report_data['truck_unit']}")
    s = report_data.get("summary",{})
    headers = ["Category","Amount"]
    rows_out = [
        ["Total Revenue",   s.get("total_revenue",0)],
        ["Loads Revenue",   s.get("loads_revenue",0)],
        ["Other Revenue",   s.get("other_revenue",0)],
        ["Driver Payments", s.get("driver_payments",0)],
        ["Fuel",            s.get("fuel",0)],
        ["Tolls",           s.get("tolls",0)],
        ["Gross Profit",    s.get("gross_profit",0)],
    ]
    return _xlsx_workbook("Gross Profit Report", meta, headers, rows_out)


def generate_gross_profit_per_load_pdf(report_data: dict, filters: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf, landscape_mode=True)
    story = []
    meta = [
        f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}",
    ]
    if report_data.get("driver_name"): meta.append(f"Driver: {report_data['driver_name']}")
    if report_data.get("truck_unit"):  meta.append(f"Truck: {report_data['truck_unit']}")
    meta.append(f"Status: {filters.get('statuses','All')}")
    meta.append(f"Billing status: {filters.get('billing_statuses','All')}")
    _header_table(story, "Gross Profit per Load Report", meta)

    headers = ["Pickup date","Delivery date","Load #","Truck #","Driver","Route","Total miles","Invoice","QP/Fac fee","Lumpers+Other","Driver Pay","Gross Profit"]
    right_cols = [6,7,8,9,10,11]
    rows_out = []
    for r in report_data.get("rows",[]):
        route = f"{r.get('pickup_city','')}, {r.get('pickup_state','')} - {r.get('delivery_city','')}, {r.get('delivery_state','')}"
        rows_out.append([
            _fmt_date(r.get("pickup_date")), _fmt_date(r.get("actual_delivery_date")),
            r.get("load_number",""), r.get("truck",""), r.get("driver",""),
            route, _fmt_num(r.get("total_miles",0)),
            _fmt_currency(r.get("rate",0)), _fmt_currency(r.get("qp_fee",0)),
            _fmt_currency(float(r.get("lumpers",0))+float(r.get("other_add_ded",0))),
            _fmt_currency(r.get("driver_pay",0)), _fmt_currency(r.get("gross_profit",0)),
        ])

    s = report_data.get("summary",{})
    totals = ["","","","","","Total:",
              _fmt_num(sum(r.get("total_miles",0) for r in report_data.get("rows",[]))),
              _fmt_currency(s.get("total_revenue",0)), "$0.00","$0.00",
              _fmt_currency(s.get("total_driver_pay",0)),
              _fmt_currency(s.get("total_gross_profit",0))]

    avail = 10.1*inch
    cw_raw = [0.65,0.65,0.5,0.6,1.1,1.4,0.6,0.7,0.6,0.7,0.7,0.7]
    t=sum(cw_raw); cw=[x/t*avail for x in cw_raw]
    _data_table(story, headers, rows_out, totals, col_widths=cw, right_align_cols=right_cols)
    doc.build(story)
    buf.seek(0)
    return buf.read()


def generate_gross_profit_per_load_xlsx(report_data: dict, filters: dict) -> bytes:
    meta = [f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}"]
    if report_data.get("driver_name"): meta.append(f"Driver: {report_data['driver_name']}")
    if report_data.get("truck_unit"):  meta.append(f"Truck: {report_data['truck_unit']}")
    headers = ["Pickup date","Delivery date","Load #","Truck #","Driver","Route","Total miles","Invoice","QP/Fac fee","Lumpers+Other","Driver Pay","Gross Profit"]
    rows_out = []
    for r in report_data.get("rows",[]):
        route = f"{r.get('pickup_city','')}, {r.get('pickup_state','')} - {r.get('delivery_city','')}, {r.get('delivery_state','')}"
        rows_out.append([_fmt_date(r.get("pickup_date")),_fmt_date(r.get("actual_delivery_date")),r.get("load_number",""),r.get("truck",""),r.get("driver",""),route,r.get("total_miles",0),r.get("rate",0),r.get("qp_fee",0),float(r.get("lumpers",0))+float(r.get("other_add_ded",0)),r.get("driver_pay",0),r.get("gross_profit",0)])
    s = report_data.get("summary",{})
    totals=["","","","","","Total",sum(r.get("total_miles",0) for r in report_data.get("rows",[])),s.get("total_revenue",0),0,0,s.get("total_driver_pay",0),s.get("total_gross_profit",0)]
    return _xlsx_workbook("Gross Profit per Load", meta, headers, rows_out, totals)


def generate_revenue_by_dispatcher_pdf(report_data: dict, filters: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf)
    story = []
    meta = [
        f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}",
        f"Dispatcher: {filters.get('dispatcher_name','All')}",
        f"Status: {filters.get('statuses','All')}",
        f"Billing status: {filters.get('billing_statuses','All')}",
    ]
    _header_table(story, "Revenue by Dispatcher Report", meta)
    headers = ["#","Dispatcher","Rate Amount","Overridden Rate"]
    rows_out = [[str(i+1), r.get("dispatcher",""), _fmt_currency(r.get("rate_amount",0)), _fmt_currency(r.get("overridden_rate",0))] for i,r in enumerate(report_data.get("rows",[]))]
    s = report_data.get("summary",{})
    totals=["","Total:", _fmt_currency(s.get("total",0)), _fmt_currency(s.get("total",0))]
    avail=7.3*inch; cw=[0.4*inch,2.5*inch,2.2*inch,2.2*inch]
    _data_table(story, headers, rows_out, totals, col_widths=cw, right_align_cols=[2,3])
    doc.build(story)
    buf.seek(0)
    return buf.read()


def generate_revenue_by_dispatcher_xlsx(report_data: dict, filters: dict) -> bytes:
    meta = [f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}", f"Dispatcher: {filters.get('dispatcher_name','All')}"]
    headers=["#","Dispatcher","Rate Amount","Overridden Rate"]
    rows_out=[[i+1,r.get("dispatcher",""),r.get("rate_amount",0),r.get("overridden_rate",0)] for i,r in enumerate(report_data.get("rows",[]))]
    s=report_data.get("summary",{})
    return _xlsx_workbook("Revenue by Dispatcher",meta,headers,rows_out,["","Total",s.get("total",0),s.get("total",0)])


def generate_payment_summary_pdf(report_data: dict, filters: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf)
    story = []
    meta = [f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}"]
    _header_table(story, "Driver Payments Summary", meta)
    headers = ["Driver Name","Payable to","Total Amount","Balance Due"]
    rows_out = [[r.get("driver_name",""),r.get("payable_to",""),_fmt_currency(r.get("total_amount",0)),_fmt_currency(r.get("balance_due",0))] for r in report_data.get("rows",[])]
    s=report_data.get("summary",{})
    totals=["","Total:",_fmt_currency(s.get("total_amount",0)),_fmt_currency(s.get("total_balance",0))]
    avail=7.3*inch; cw=[2.0*inch,2.0*inch,1.65*inch,1.65*inch]
    _data_table(story,headers,rows_out,totals,col_widths=cw,right_align_cols=[2,3])
    doc.build(story)
    buf.seek(0)
    return buf.read()


def generate_payment_summary_xlsx(report_data: dict, filters: dict) -> bytes:
    meta=[f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}"]
    headers=["Driver Name","Payable to","Total Amount","Balance Due"]
    rows_out=[[r.get("driver_name",""),r.get("payable_to",""),r.get("total_amount",0),r.get("balance_due",0)] for r in report_data.get("rows",[])]
    s=report_data.get("summary",{})
    return _xlsx_workbook("Driver Payments Summary",meta,headers,rows_out,["","Total",s.get("total_amount",0),s.get("total_balance",0)])


def generate_expenses_pdf(report_data: dict, filters: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf)
    story = []
    meta = [f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}", f"Category: {filters.get('category','All')}"]
    _header_table(story,"Expenses Report",meta)
    headers=["#","Category","Amount"]
    rows_out=[[str(i+1),r.get("category",""),_fmt_currency(r.get("amount",0))] for i,r in enumerate(report_data.get("rows",[]))]
    s=report_data.get("summary",{})
    totals=["","Total:",_fmt_currency(s.get("total",0))]
    avail=7.3*inch; cw=[0.4*inch,4.5*inch,2.4*inch]
    _data_table(story,headers,rows_out,totals,col_widths=cw,right_align_cols=[2])
    doc.build(story)
    buf.seek(0)
    return buf.read()


def generate_expenses_xlsx(report_data: dict, filters: dict) -> bytes:
    meta=[f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}", f"Category: {filters.get('category','All')}"]
    headers=["#","Category","Amount"]
    rows_out=[[i+1,r.get("category",""),r.get("amount",0)] for i,r in enumerate(report_data.get("rows",[]))]
    s=report_data.get("summary",{})
    return _xlsx_workbook("Expenses Report",meta,headers,rows_out,["","Total",s.get("total",0)])


def generate_profit_loss_pdf(report_data: dict, filters: dict) -> bytes:
    buf = io.BytesIO()
    doc = _pdf_doc(buf)
    story = []
    meta = [f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}"]
    if report_data.get("driver_name"): meta.append(f"Driver: {report_data['driver_name']}")
    if report_data.get("truck_unit"):  meta.append(f"Truck: {report_data['truck_unit']}")
    _header_table(story,"Profit & Loss Report",meta)
    s=report_data.get("summary",{})
    headers=["","Amount"]
    rows_out=[
        ["Total Revenue",  _fmt_currency(s.get("total_revenue",0))],
        ["Loads Revenue",  _fmt_currency(s.get("loads_revenue",0))],
        ["Other Revenue",  _fmt_currency(s.get("other_revenue",0))],
        ["Driver Payments",_fmt_currency(s.get("driver_payments",0))],
        ["Fuel",           _fmt_currency(s.get("fuel",0))],
        ["Tolls",          _fmt_currency(s.get("tolls",0))],
        ["Expenses",       _fmt_currency(s.get("expenses",0))],
        ["Gross Profit",   _fmt_currency(s.get("gross_profit",0))],
        ["Net Profit",     _fmt_currency(s.get("net_profit",0))],
    ]
    cw=[5.0*inch,2.3*inch]
    _data_table(story,headers,rows_out,col_widths=cw,right_align_cols=[1])
    doc.build(story)
    buf.seek(0)
    return buf.read()


def generate_profit_loss_xlsx(report_data: dict, filters: dict) -> bytes:
    meta=[f"Dates range: {_fmt_date(report_data.get('date_from'))} {_fmt_date(report_data.get('date_to'))}"]
    if report_data.get("driver_name"): meta.append(f"Driver: {report_data['driver_name']}")
    if report_data.get("truck_unit"):  meta.append(f"Truck: {report_data['truck_unit']}")
    s=report_data.get("summary",{})
    headers=["Category","Amount"]
    rows_out=[["Total Revenue",s.get("total_revenue",0)],["Loads Revenue",s.get("loads_revenue",0)],["Other Revenue",s.get("other_revenue",0)],["Driver Payments",s.get("driver_payments",0)],["Fuel",s.get("fuel",0)],["Tolls",s.get("tolls",0)],["Expenses",s.get("expenses",0)],["Gross Profit",s.get("gross_profit",0)],["Net Profit",s.get("net_profit",0)]]
    return _xlsx_workbook("Profit & Loss Report",meta,headers,rows_out)
